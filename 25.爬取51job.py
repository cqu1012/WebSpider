# -*- coding: utf-8 -*-
#https://search.51job.com/list/090200%252C040000,000000,0000,00,9,99,python%25E5%25BC%2580%25E5%258F%2591%25E5%25B7%25A5%25E7%25A8%258B%25E5%25B8%2588%2B,2,1.html?lang=c&stype=1&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&lonlat=0%2C0&radius=-1&ord_field=0&confirmdate=9&fromType=&dibiaoid=0&address=&line=&specialarea=00&from=&welfare=
#https://search.51job.com/list/090200%252C040000,000000,0000,00,9,99,python%25E5%25BC%2580%25E5%258F%2591%25E5%25B7%25A5%25E7%25A8%258B%25E5%25B8%2588%2B,2,2.html?lang=c&stype=1&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&lonlat=0%2C0&radius=-1&ord_field=0&confirmdate=9&fromType=&dibiaoid=0&address=&line=&specialarea=00&from=&welfare=
from lxml import etree
from fake_useragent import UserAgent
import requests
from time import sleep
import openpyxl

headers = {'User-Agent':UserAgent().random}
def getHtml(url):
    rsp = requests.get(url,headers=headers)
    html = etree.HTML(rsp.text)
    postionList = html.xpath("//div[@class='el']/p/span/a/text()")
    urlList = html.xpath("//div[@class='el']/p/span/a/@href")
    return postionList,urlList

def getDtailMsg(ws,url_list):
    '''
    :param url_list: url列表
    :return: 返回职位，公司，薪水，要求
    '''
    for url in url_list:
        rsp = requests.get(url,headers=headers)
        html = etree.HTML(rsp.text)
        position = html.xpath('//h1/text()')[0].strip()
        company = html.xpath("//p[@class='cname']/a[@class='catn']/text()")[0].strip()
        salary = html.xpath("//div[@class='cn']/strong/text()")
        if len(salary) == 0:
            slary_num = '面谈'
        else:
            slary_num = salary[0]
        #职位信息为1个list
        postionInfo = html.xpath("//div[@class='bmsg job_msg inbox']/p/text()")
        if len(postionInfo) == 0:
            postionInfo1 =html.xpath("//div[@class='bmsg job_msg inbox']/text()")
        else:
            postionInfo1 = postionInfo
        dictmsg={'position': position, 'company': company, 'salary': slary_num, 'positionInfo': postionInfo1,}
        ws.append([dictmsg.get('position'), dictmsg.get('company'), str(dictmsg.get('salary')), str(dictmsg.get('positionInfo')).strip()])
        sleep(2)
    return True

def main():
    base_url = "https://search.51job.com/list/090200%252C040000,000000,0000,00,9,99,python%25E5%25BC%2580%25E5%258F%2591%25E5%25B7%25A5%25E7%25A8%258B%25E5%25B8%2588%2B,2,{}.html?lang=c&stype=1&postchannel=0000&workyear=99&cotype=99&degreefrom=99&jobterm=99&companysize=99&lonlat=0%2C0&radius=-1&ord_field=0&confirmdate=9&fromType=&dibiaoid=0&address=&line=&specialarea=00&from=&welfare="
    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'title1'
    ws['A1'].value = 'position'
    ws['B1'].value = 'company'
    ws['C1'].value = 'salary'
    ws['D1'].value = 'positionInfo'
    for i in range(1):
        url = base_url.format(i+1)
        positionList,urlList=getHtml(url)
        getDtailMsg(ws, urlList)
        sleep(3)
    wb.save('Python工程师求职信息.xlsx')

if __name__ =='__main__':
    main()